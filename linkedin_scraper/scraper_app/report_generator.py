import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import pandas as pd


# ── Colour palette ─────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
ACCENT      = "00B0F0"
WHITE       = "FFFFFF"
LIGHT_GREY  = "F2F2F2"
DARK_GREY   = "404040"
GREEN       = "70AD47"
ORANGE      = "ED7D31"


def _border(style="thin"):
    s = Side(border_style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    return c


def _data_cell(ws, row, col, value, bg=WHITE, bold=False, align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=10, color=DARK_GREY)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border()
    if num_fmt:
        c.number_format = num_fmt
    return c


# ── Sheet 1 – Raw Data ─────────────────────────────────────────────────────────
def _build_raw_sheet(wb, df):
    ws = wb.active
    ws.title = "📋 Raw Data"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "LinkedIn Job Scraper — Raw Data"
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    c.font = Font(name="Arial", size=9, color="888888")
    c.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = ["#", "Job Title", "Company", "Location", "Date Posted", "Scraped At", "Job Link"]
    widths  = [5,   40,          28,         22,         14,            18,            55]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws, 3, i, h, bg=MID_BLUE)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 22

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=4):
        bg = WHITE if r_idx % 2 == 0 else LIGHT_GREY
        _data_cell(ws, r_idx, 1, r_idx - 3,          bg=bg, align="center")
        _data_cell(ws, r_idx, 2, row.title,           bg=bg)
        _data_cell(ws, r_idx, 3, row.company,         bg=bg)
        _data_cell(ws, r_idx, 4, row.location,        bg=bg)
        _data_cell(ws, r_idx, 5, row.date_posted,     bg=bg, align="center")
        _data_cell(ws, r_idx, 6, row.scraped_at[:16], bg=bg, align="center")

        # hyperlink for job link
        cell = ws.cell(row=r_idx, column=7)
        link = str(row.job_link)
        if link.startswith("http"):
            cell.value = "View Job"
            cell.hyperlink = link
            cell.font = Font(name="Arial", size=10, color=ACCENT, underline="single")
        else:
            cell.value = link
            cell.font = Font(name="Arial", size=10, color=DARK_GREY)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"


# ── Sheet 2 – Summary Stats ────────────────────────────────────────────────────
def _build_summary_sheet(wb, df, keyword):
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Job Market Summary — '{keyword}'"
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # KPI boxes  (row 3-6)
    kpis = [
        ("Total Jobs",        len(df),                              MID_BLUE),
        ("Unique Companies",  df["company"].nunique(),               GREEN),
        ("Unique Locations",  df["location"].nunique(),              ORANGE),
        ("Latest Scrape",     datetime.now().strftime("%d %b %Y"),  DARK_BLUE),
    ]
    for i, (label, value, color) in enumerate(kpis, 1):
        col = (i - 1) * 2 + 1          # cols 1,3,5,7
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)

        # label
        lc = ws.cell(row=3, column=col, value=label)
        lc.font = Font(name="Arial", bold=False, size=9, color=WHITE)
        lc.fill = PatternFill("solid", fgColor=color)
        lc.alignment = Alignment(horizontal="center", vertical="center")

        # value
        vc = ws.cell(row=4, column=col, value=value)
        vc.font = Font(name="Arial", bold=True, size=20, color=WHITE)
        vc.fill = PatternFill("solid", fgColor=color)
        vc.alignment = Alignment(horizontal="center", vertical="center")

        # spacer
        sc = ws.cell(row=5, column=col, value="")
        sc.fill = PatternFill("solid", fgColor=color)

        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 38
        ws.row_dimensions[5].height = 8

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Top 10 Companies table  (row 8)
    top_companies = df["company"].value_counts().head(10).reset_index()
    top_companies.columns = ["Company", "Job Count"]

    _header_cell(ws, 7, 1, "Top 10 Companies by Listings", bg=MID_BLUE, size=10)
    ws.merge_cells("A7:C7")
    ws.row_dimensions[7].height = 20

    for col_i, h in enumerate(["Rank", "Company", "Jobs"], 1):
        _header_cell(ws, 8, col_i, h, bg=DARK_BLUE, size=10)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10

    for idx, row in enumerate(top_companies.itertuples(index=False), start=9):
        bg = WHITE if idx % 2 == 0 else LIGHT_GREY
        _data_cell(ws, idx, 1, idx - 8,         bg=bg, align="center")
        _data_cell(ws, idx, 2, row.Company,      bg=bg)
        _data_cell(ws, idx, 3, row[1],           bg=bg, align="center", bold=True)

    # Top 10 Locations table  (same rows, cols 5-7)
    top_locs = df["location"].value_counts().head(10).reset_index()
    top_locs.columns = ["Location", "Job Count"]

    _header_cell(ws, 7, 5, "Top 10 Locations by Listings", bg=MID_BLUE, size=10)
    ws.merge_cells("E7:G7")

    for col_i, h in enumerate(["Rank", "Location", "Jobs"], 5):
        _header_cell(ws, 8, col_i, h, bg=DARK_BLUE, size=10)
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 10

    for idx, row in enumerate(top_locs.itertuples(index=False), start=9):
        bg = WHITE if idx % 2 == 0 else LIGHT_GREY
        _data_cell(ws, idx, 5, idx - 8,         bg=bg, align="center")
        _data_cell(ws, idx, 6, row.Location,     bg=bg)
        _data_cell(ws, idx, 7, row[1],           bg=bg, align="center", bold=True)


# ── Sheet 3 – Charts ───────────────────────────────────────────────────────────
def _build_charts_sheet(wb, df):
    ws = wb.create_sheet("📈 Charts")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value = "Visual Analytics"
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Data table 1: Companies (hidden cols A-B used by chart) ────────────────
    top_co = df["company"].value_counts().head(8).reset_index()
    top_co.columns = ["Company", "Count"]

    ws["A3"] = "Company"
    ws["B3"] = "Count"
    for i, row in enumerate(top_co.itertuples(index=False), start=4):
        ws.cell(row=i, column=1, value=row.Company)
        ws.cell(row=i, column=2, value=row[1])

    # Bar chart – Top Companies
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top Companies Hiring"
    bar.y_axis.title = "Company"
    bar.x_axis.title = "Job Listings"
    bar.style = 10
    bar.width = 18
    bar.height = 12

    data_ref  = Reference(ws, min_col=2, min_row=3, max_row=3 + len(top_co))
    cats_ref  = Reference(ws, min_col=1, min_row=4, max_row=3 + len(top_co))
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = MID_BLUE
    ws.add_chart(bar, "D3")

    # ── Data table 2: Locations (cols A-B continued) ───────────────────────────
    top_loc = df["location"].value_counts().head(7).reset_index()
    top_loc.columns = ["Location", "Count"]

    row_offset = 4 + len(top_co) + 2
    ws.cell(row=row_offset,   column=1, value="Location")
    ws.cell(row=row_offset,   column=2, value="Count")
    for i, row in enumerate(top_loc.itertuples(index=False), start=row_offset + 1):
        ws.cell(row=i, column=1, value=row.Location)
        ws.cell(row=i, column=2, value=row[1])

    # Pie chart – Location distribution
    pie = PieChart()
    pie.title = "Jobs by Location"
    pie.style = 10
    pie.width  = 14
    pie.height = 12

    pie_data = Reference(ws, min_col=2, min_row=row_offset, max_row=row_offset + len(top_loc))
    pie_cats = Reference(ws, min_col=1, min_row=row_offset + 1, max_row=row_offset + len(top_loc))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.dataLabels = None
    # colour slices
    colors = [MID_BLUE, ACCENT, GREEN, ORANGE, "FF0000", "9B59B6", "F39C12"]
    for j, color in enumerate(colors[:len(top_loc)]):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)

    ws.add_chart(pie, "D22")

    # hide the raw helper cols visually
    for col in ["A", "B", "C"]:
        ws.column_dimensions[col].width = 0.5


# ── Master function ────────────────────────────────────────────────────────────
def generate_report(rows: list[dict], keyword: str, output_path: str) -> str:
    df = pd.DataFrame(rows)
    if df.empty:
        raise ValueError("No job data to generate report from.")

    # Clean up
    df["company"]     = df["company"].fillna("Unknown").str.strip()
    df["location"]    = df["location"].fillna("Unknown").str.strip()
    df["title"]       = df["title"].fillna("").str.strip()
    df["date_posted"] = df["date_posted"].fillna("").str.strip()
    df["scraped_at"]  = df["scraped_at"].fillna("").str.strip()
    df["job_link"]    = df["job_link"].fillna("").str.strip()

    wb = Workbook()
    _build_raw_sheet(wb, df)
    _build_summary_sheet(wb, df, keyword)
    _build_charts_sheet(wb, df)

    wb.save(output_path)
    return output_path
